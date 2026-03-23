#!/usr/bin/env python3
"""
build_qbank.py — AI Readiness Assessment Question Bank Generator v2.0
Author:  Gagandeep Singh | AI Practice
Version: 2.0.0

Generates public/ai-question-bank.xlsx with:
  - 25 original organisational functions (enhanced with cross-cutting questions)
  - 2 NEW standalone functions: data_readiness, ai_literacy
  - 21-column schema (unchanged from v1)
  - DQ-tagged questions in data_readiness carry 2x points (points_yes=4)

Schema columns (21):
  function_id, function_name, domain, competency_level, question_id,
  question_text, answer_type, branch_trigger, follow_up_text,
  dim_DQ, dim_TR, dim_TS, dim_GE, dim_CR, dim_VR,
  points_yes, points_partial, points_no,
  industry_overlay, ai_use_case_hint, regulatory_ref
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl import Workbook
import os

# ─── COLUMN HEADERS ──────────────────────────────────────────────────────────
COLS = [
    "function_id", "function_name", "domain", "competency_level",
    "question_id", "question_text", "answer_type",
    "branch_trigger", "follow_up_text",
    "dim_DQ", "dim_TR", "dim_TS", "dim_GE", "dim_CR", "dim_VR",
    "points_yes", "points_partial", "points_no",
    "industry_overlay", "ai_use_case_hint", "regulatory_ref"
]

def row(function_id, function_name, domain, level, q_id, question_text,
        answer_type="YPN", branch_trigger="", follow_up="",
        dq=0, tr=0, ts=0, ge=0, cr=0, vr=0,
        pts_yes=2, pts_partial=1, pts_no=0,
        industry="ALL", use_case_hint="", reg_ref=""):
    return {
        "function_id": function_id, "function_name": function_name,
        "domain": domain, "competency_level": level,
        "question_id": q_id, "question_text": question_text,
        "answer_type": answer_type,
        "branch_trigger": branch_trigger, "follow_up_text": follow_up,
        "dim_DQ": dq, "dim_TR": tr, "dim_TS": ts, "dim_GE": ge, "dim_CR": cr, "dim_VR": vr,
        "points_yes": pts_yes, "points_partial": pts_partial, "points_no": pts_no,
        "industry_overlay": industry, "ai_use_case_hint": use_case_hint,
        "regulatory_ref": reg_ref
    }

# DQ 2x points helper for data_readiness DQ questions
def dq2x_row(function_id, function_name, domain, level, q_id, question_text,
              answer_type="YPN", branch_trigger="", follow_up="",
              dq=1, tr=0, ts=0, ge=0, cr=0, vr=0,
              pts_yes=None, pts_partial=None,
              industry="ALL", use_case_hint="", reg_ref=""):
    """DQ-tagged data_readiness questions get 2x points (4/2/0). pts_yes/pts_partial override the auto-calc."""
    y = pts_yes if pts_yes is not None else (4 if dq else 2)
    p = pts_partial if pts_partial is not None else (2 if dq else 1)
    return row(function_id, function_name, domain, level, q_id, question_text,
               answer_type, branch_trigger, follow_up,
               dq, tr, ts, ge, cr, vr,
               pts_yes=y, pts_partial=p, pts_no=0,
               industry=industry, use_case_hint=use_case_hint, reg_ref=reg_ref)

ALL_ROWS = []

# ═══════════════════════════════════════════════════════════════════════════════
# NEW FUNCTION 1: DATA READINESS
# ═══════════════════════════════════════════════════════════════════════════════
FN = "data_readiness"
FN_NAME = "Data Readiness"
DOM = "Technology"

# EXPLORING (6 questions)
ALL_ROWS += [
    dq2x_row(FN, FN_NAME, DOM, "Exploring", "DE-E-01",
        "Does your organisation have a data inventory or data catalogue that documents what data assets exist, where they are stored, and who owns them?",
        branch_trigger="No", follow_up="What is the primary barrier — no ownership assigned, no tooling, or data spread across too many systems?",
        dq=1, ge=1, use_case_hint="Data governance foundation", reg_ref="GDPR Art.30"),
    dq2x_row(FN, FN_NAME, DOM, "Exploring", "DE-E-02",
        "What proportion of your business-critical data is stored in structured, machine-readable formats (databases, data warehouses) versus unstructured formats (emails, PDFs, spreadsheets, paper)?",
        dq=1, tr=1, use_case_hint="Structured data readiness for ML", reg_ref=""),
    dq2x_row(FN, FN_NAME, DOM, "Exploring", "DE-E-03",
        "Is there a named Data Owner or Data Steward for each major data domain (customer, financial, operational, HR) who is accountable for data quality?",
        dq=1, ge=1, use_case_hint="Data stewardship", reg_ref="GDPR Art.5"),
    dq2x_row(FN, FN_NAME, DOM, "Exploring", "DE-E-04",
        "Has a formal data quality assessment been conducted in the last 24 months — measuring accuracy, completeness, consistency, and timeliness of key datasets?",
        dq=1, use_case_hint="Data quality baseline", reg_ref=""),
    dq2x_row(FN, FN_NAME, DOM, "Exploring", "DE-E-05",
        "Is your organisation subject to data residency restrictions (e.g. GDPR requiring EU data to stay in EU) that would constrain where AI models can be trained or where data can be sent for processing?",
        ge=1, tr=1, pts_yes=2, pts_partial=1,
        use_case_hint="AI data sovereignty planning", reg_ref="GDPR Art.44-49, EU AI Act"),
    dq2x_row(FN, FN_NAME, DOM, "Exploring", "DE-E-06",
        "Does your organisation have a documented data retention and deletion policy that is actively enforced — relevant to AI training data governance under GDPR Art.5?",
        dq=1, ge=1, use_case_hint="Training data governance", reg_ref="GDPR Art.5(1)(e)"),
]

# PILOTING (7 questions)
ALL_ROWS += [
    dq2x_row(FN, FN_NAME, DOM, "Piloting", "DE-P-01",
        "Is there a central data platform (data warehouse, data lake, or lakehouse such as Snowflake, Databricks, BigQuery) that makes data accessible across the organisation for analytics and AI use cases?",
        dq=1, tr=1, use_case_hint="Unified AI data platform", reg_ref=""),
    dq2x_row(FN, FN_NAME, DOM, "Piloting", "DE-P-02",
        "Are data pipelines in place that automatically move, clean, and transform data from source systems into the central platform — with documented lineage showing exactly how each dataset was produced?",
        dq=1, tr=1, use_case_hint="ML-ready data pipelines", reg_ref="EU AI Act Art.10"),
    dq2x_row(FN, FN_NAME, DOM, "Piloting", "DE-P-03",
        "Has your organisation identified its 'crown jewel' datasets — proprietary data assets that competitors do not have access to and that could form the basis of AI competitive advantage?",
        dq=1, vr=1, use_case_hint="Proprietary AI data strategy", reg_ref=""),
    dq2x_row(FN, FN_NAME, DOM, "Piloting", "DE-P-04",
        "Is data quality monitored continuously — with automated alerts when data quality metrics (null rates, duplicate rates, schema changes) breach defined thresholds?",
        dq=1, tr=1, use_case_hint="Automated data quality monitoring", reg_ref=""),
    dq2x_row(FN, FN_NAME, DOM, "Piloting", "DE-P-05",
        "Are Master Data Management (MDM) standards in place for key entities (customer, product, supplier, employee) — ensuring a single authoritative definition across all systems?",
        dq=1, ge=1, use_case_hint="MDM for AI training consistency", reg_ref=""),
    row(FN, FN_NAME, DOM, "Piloting", "DE-P-06",
        "Has a Data Privacy Impact Assessment (DPIA) been completed for any AI use case that processes personal data — as required by GDPR Article 35?",
        ge=1, use_case_hint="DPIA for AI", reg_ref="GDPR Art.35"),
    dq2x_row(FN, FN_NAME, DOM, "Piloting", "DE-P-07",
        "Is there a formal data access governance process — defining who can access which datasets, with audit logs of all data access for regulatory compliance?",
        dq=1, ge=1, use_case_hint="Data access controls for AI", reg_ref="GDPR Art.32, EU AI Act Art.10"),
]

# SCALING (5 questions)
ALL_ROWS += [
    dq2x_row(FN, FN_NAME, DOM, "Scaling", "DE-S-01",
        "Does the organisation have a proprietary training data strategy — actively curating, labelling, and enriching datasets that will be used to fine-tune AI models for competitive advantage unavailable to rivals?",
        dq=1, vr=1, use_case_hint="Proprietary fine-tuning dataset", reg_ref="EU AI Act Art.10"),
    dq2x_row(FN, FN_NAME, DOM, "Scaling", "DE-S-02",
        "Is synthetic data generation used to augment training datasets where real data is scarce, sensitive, or insufficiently diverse?",
        dq=1, tr=1, use_case_hint="Synthetic data for AI training", reg_ref="GDPR Art.89"),
    dq2x_row(FN, FN_NAME, DOM, "Scaling", "DE-S-03",
        "Are AI model training datasets versioned and reproducible — so any production model can be retrained from a known, archived dataset state for regulatory audit purposes?",
        dq=1, ge=1, use_case_hint="Dataset versioning for model audit", reg_ref="EU AI Act Art.12"),
    row(FN, FN_NAME, DOM, "Scaling", "DE-S-04",
        "Is there a Federated Learning or Privacy-Preserving ML strategy for use cases where data cannot be centralised due to privacy or sovereignty constraints?",
        ge=1, tr=1, use_case_hint="Federated learning deployment", reg_ref="GDPR Art.25, EU AI Act"),
    dq2x_row(FN, FN_NAME, DOM, "Scaling", "DE-S-05",
        "Does the organisation publish a Data Strategy that is board-ratified, includes AI data objectives, and has named executive ownership and a funded multi-year roadmap?",
        dq=1, ge=1, cr=1, use_case_hint="Enterprise AI data strategy", reg_ref="EU AI Act Art.9"),
]

# ═══════════════════════════════════════════════════════════════════════════════
# NEW FUNCTION 2: AI LITERACY & WORKFORCE DEVELOPMENT
# ═══════════════════════════════════════════════════════════════════════════════
FN = "ai_literacy"
FN_NAME = "AI Literacy & Training"
DOM = "Business"

# EXPLORING (5 questions)
ALL_ROWS += [
    row(FN, FN_NAME, DOM, "Exploring", "AL-E-01",
        "Has your organisation assessed the current baseline AI literacy of its workforce — understanding what proportion of employees know what AI is, can describe its risks, and have used at least one AI tool in their work?",
        ts=1, cr=1, use_case_hint="AI literacy baseline assessment", reg_ref=""),
    row(FN, FN_NAME, DOM, "Exploring", "AL-E-02",
        "Do employees have clear guidance on what AI tools they are permitted to use for work purposes — including which tools are approved for handling confidential, personal, or regulated data?",
        ge=1, ts=1, use_case_hint="AI acceptable use policy", reg_ref="EU AI Act Art.4, GDPR Art.32"),
    row(FN, FN_NAME, DOM, "Exploring", "AL-E-03",
        "Has leadership (C-suite and senior management) received any structured AI education — covering AI capabilities, limitations, governance obligations, and strategic implications?",
        ts=1, cr=1, use_case_hint="Executive AI education", reg_ref="EU AI Act Art.4"),
    row(FN, FN_NAME, DOM, "Exploring", "AL-E-04",
        "Is AI included in the onboarding programme for new employees — ensuring every new joiner understands the organisation's AI policies, approved tools, and responsible use expectations?",
        ts=1, ge=1, use_case_hint="AI onboarding standard", reg_ref="EU AI Act Art.4"),
    row(FN, FN_NAME, DOM, "Exploring", "AL-E-05",
        "Has the organisation identified which employee roles are most exposed to AI-driven change — and communicated transparently with those employees about what AI means for their work?",
        cr=1, ts=1, use_case_hint="Workforce AI impact mapping", reg_ref=""),
]

# PILOTING (6 questions)
ALL_ROWS += [
    row(FN, FN_NAME, DOM, "Piloting", "AL-P-01",
        "Is there a structured AI upskilling programme available to all employees — covering AI fundamentals, prompt engineering, AI tool proficiency, and responsible AI use?",
        branch_trigger="Partial", follow_up="What proportion of the workforce has completed any AI training in the last 12 months?",
        ts=1, use_case_hint="Enterprise AI upskilling programme", reg_ref="EU AI Act Art.4"),
    row(FN, FN_NAME, DOM, "Piloting", "AL-P-02",
        "Do data and analytics roles in the organisation have mandatory AI and ML skills requirements — and is there a competency framework defining what 'AI-capable' means for each role family?",
        ts=1, dq=1, use_case_hint="AI competency framework", reg_ref=""),
    row(FN, FN_NAME, DOM, "Piloting", "AL-P-03",
        "Is AI literacy tracked as a measurable KPI — with a baseline score, a target, and progress reported to the CHRO or board?",
        ts=1, vr=1, use_case_hint="AI literacy KPI dashboard", reg_ref=""),
    row(FN, FN_NAME, DOM, "Piloting", "AL-P-04",
        "Are there dedicated AI Champions or AI Ambassadors in each business function — advocates trained to support colleagues in AI adoption and surface grassroots AI use cases?",
        cr=1, ts=1, use_case_hint="AI champion network", reg_ref=""),
    row(FN, FN_NAME, DOM, "Piloting", "AL-P-05",
        "Does the L&D function use AI-powered tools to personalise learning pathways for employees — demonstrating AI use in the function that most advocates for AI skills?",
        ts=1, tr=1, vr=1, use_case_hint="AI-powered personalised learning", reg_ref=""),
    row(FN, FN_NAME, DOM, "Piloting", "AL-P-06",
        "Has the organisation mapped skills gaps relative to its AI ambitions — identifying specific technical, analytical, and leadership capabilities that need to be built or hired?",
        ts=1, cr=1, use_case_hint="AI skills gap analysis", reg_ref=""),
]

# SCALING (4 questions)
ALL_ROWS += [
    row(FN, FN_NAME, DOM, "Scaling", "AL-S-01",
        "Does the organisation have a multi-year AI talent strategy — covering hiring, reskilling, retention of AI-capable staff, and partnerships with universities or training providers?",
        ts=1, cr=1, use_case_hint="Strategic AI talent pipeline", reg_ref=""),
    row(FN, FN_NAME, DOM, "Scaling", "AL-S-02",
        "Is there a board-level commitment to AI workforce investment — with a named budget for AI skills development and progress reported to the Remuneration Committee?",
        ts=1, cr=1, vr=1, use_case_hint="Board AI workforce investment", reg_ref="EU AI Act Art.9"),
    row(FN, FN_NAME, DOM, "Scaling", "AL-S-03",
        "Has the organisation achieved measurable AI productivity gains attributable to employee AI literacy — with before/after metrics showing time saved, quality improved, or cost reduced through AI tool adoption?",
        ts=1, vr=1, use_case_hint="AI literacy ROI measurement", reg_ref=""),
    row(FN, FN_NAME, DOM, "Scaling", "AL-S-04",
        "Does the organisation contribute to external AI talent ecosystems — through apprenticeships, graduate programmes, academic partnerships, or industry working groups — building AI talent pipeline beyond just internal training?",
        ts=1, cr=1, use_case_hint="External AI talent ecosystem", reg_ref=""),
]

# ═══════════════════════════════════════════════════════════════════════════════
# ORIGINAL 25 FUNCTIONS — with full question sets + 2 cross-cutting questions each
# ═══════════════════════════════════════════════════════════════════════════════
# Helper to build cross-cutting questions for each function
def cc(fn_id, fn_name, dom, data_q, train_q):
    return [
        row(fn_id, fn_name, dom, "Piloting", f"{fn_id.upper()[:6]}-P-DATA",
            data_q, dq=1,
            use_case_hint="Function data readiness for AI", reg_ref="EU AI Act Art.10"),
        row(fn_id, fn_name, dom, "Piloting", f"{fn_id.upper()[:6]}-P-TRAIN",
            train_q, ts=1, ge=1,
            use_case_hint="Function AI tool literacy", reg_ref="EU AI Act Art.4"),
    ]

# ─── FINANCE & ACCOUNTING ─────────────────────────────────────────────────────
FN, FN_NAME, DOM = "finance", "Finance & Accounting", "Business"
ALL_ROWS += [
    row(FN, FN_NAME, DOM, "Exploring", "FIN-E-01",
        "Has the Finance function explored any AI or ML tools for tasks such as expense categorisation, invoice processing, fraud detection, or financial forecasting?",
        cr=1, vr=1, use_case_hint="AI exploration in finance"),
    row(FN, FN_NAME, DOM, "Exploring", "FIN-E-02",
        "Does the Finance team have access to clean, consistent, and timely financial data from all entities and systems, consolidated in a single accessible platform?",
        dq=1, use_case_hint="Finance data consolidation"),
    row(FN, FN_NAME, DOM, "Exploring", "FIN-E-03",
        "Is there leadership sponsorship within Finance for AI-driven process improvement — a CFO or Finance Director who champions automation and AI adoption?",
        cr=1, use_case_hint="CFO AI sponsorship"),
    row(FN, FN_NAME, DOM, "Piloting", "FIN-P-01",
        "Has the Finance function piloted any AI tool for accounts payable automation, intelligent invoice matching, or automated journal entry generation?",
        tr=1, vr=1, use_case_hint="AP automation AI pilot"),
    row(FN, FN_NAME, DOM, "Piloting", "FIN-P-02",
        "Is AI used in financial planning & analysis (FP&A) — for scenario modelling, variance commentary generation, or rolling forecast accuracy improvement?",
        tr=1, vr=1, dq=1, use_case_hint="AI FP&A forecasting"),
    row(FN, FN_NAME, DOM, "Piloting", "FIN-P-03",
        "Has the Finance function implemented or piloted AI-powered fraud detection — monitoring transactions in real time against behavioural baselines?",
        tr=1, ge=1, vr=1, use_case_hint="Real-time fraud detection AI", reg_ref="FCA / SOX"),
    row(FN, FN_NAME, DOM, "Scaling", "FIN-S-01",
        "Is AI embedded in the financial close process — automating reconciliations, flagging anomalies, and reducing close cycle time below industry average?",
        tr=1, vr=1, dq=1, use_case_hint="AI-powered financial close"),
    row(FN, FN_NAME, DOM, "Scaling", "FIN-S-02",
        "Are AI-generated financial insights and commentary reviewed by qualified humans before publication — with a clear audit trail of AI vs human contributions for regulatory compliance?",
        ge=1, ts=1, use_case_hint="Human-in-loop financial AI", reg_ref="SOX, IFRS"),
]
ALL_ROWS += cc(FN, FN_NAME, DOM,
    "Does the Finance team have reliable, accessible, clean data — including multi-year historical financials, ERP transaction data, and GL master data — sufficient to train or fine-tune an AI forecasting or reconciliation model? Is data quality monitored, and are known data gaps on a remediation roadmap?",
    "Has the Finance team been trained on AI-assisted reconciliation and forecasting tools, including how to detect when AI-generated journal entries or variance commentary contains errors, and what escalation procedures exist when AI outputs are materially wrong?")

# ─── HUMAN RESOURCES ─────────────────────────────────────────────────────────
FN, FN_NAME, DOM = "hr", "Human Resources", "Business"
ALL_ROWS += [
    row(FN, FN_NAME, DOM, "Exploring", "HR-E-01",
        "Has the HR function explored the use of AI in any talent process — such as CV screening, candidate matching, interview scheduling, or engagement survey analysis?",
        cr=1, vr=1, use_case_hint="AI in talent acquisition"),
    row(FN, FN_NAME, DOM, "Exploring", "HR-E-02",
        "Does HR have access to clean employee data (performance, tenure, skills, pay, absence) in a single accessible system that could be used for workforce analytics?",
        dq=1, use_case_hint="HR people analytics data"),
    row(FN, FN_NAME, DOM, "Exploring", "HR-E-03",
        "Is the CHRO or HR Director aware of the EU AI Act's specific requirements for AI used in recruitment, performance evaluation, and employee monitoring — and has legal review been conducted?",
        ge=1, use_case_hint="HR AI regulatory awareness", reg_ref="EU AI Act Annex III, GDPR Art.22"),
    row(FN, FN_NAME, DOM, "Piloting", "HR-P-01",
        "Is AI used in the recruitment process — for job description optimisation, CV ranking, or candidate engagement — with documented bias testing and adverse impact analysis?",
        tr=1, ge=1, vr=1, use_case_hint="Bias-tested AI recruitment", reg_ref="EU AI Act Annex III"),
    row(FN, FN_NAME, DOM, "Piloting", "HR-P-02",
        "Does HR use predictive attrition modelling — identifying flight-risk employees and enabling proactive retention interventions before resignation?",
        tr=1, dq=1, vr=1, use_case_hint="Predictive attrition AI"),
    row(FN, FN_NAME, DOM, "Scaling", "HR-S-01",
        "Is AI embedded across the full employee lifecycle — from AI-assisted onboarding personalisation to AI-driven learning recommendations and AI-powered exit interview analysis?",
        tr=1, ts=1, vr=1, use_case_hint="Full-lifecycle HR AI"),
    row(FN, FN_NAME, DOM, "Scaling", "HR-S-02",
        "Does HR use AI to enable hyper-personalised employee experiences — tailored benefits, career path recommendations, and wellbeing nudges based on individual data?",
        tr=1, dq=1, ge=1, vr=1, use_case_hint="Personalised employee AI", reg_ref="GDPR Art.22"),
]
ALL_ROWS += cc(FN, FN_NAME, DOM,
    "Does the HR team have reliable, accessible, clean people data — including structured performance ratings, skills data, compensation history, and absence records — sufficient to train a predictive attrition or skills-matching AI model? Is data quality monitored, and are known data gaps on a remediation roadmap?",
    "Have HR staff who use AI screening or performance tools been trained on bias indicators, adverse impact ratios, and how to override AI recommendations when warranted — including what documentation is required when a human overrides an AI decision?")

# ─── SALES & REVENUE ─────────────────────────────────────────────────────────
FN, FN_NAME, DOM = "sales", "Sales & Revenue", "Business"
ALL_ROWS += [
    row(FN, FN_NAME, DOM, "Exploring", "SAL-E-01",
        "Has the Sales function explored or used any AI tools for lead scoring, pipeline forecasting, or customer propensity modelling?",
        cr=1, vr=1, use_case_hint="AI lead scoring exploration"),
    row(FN, FN_NAME, DOM, "Exploring", "SAL-E-02",
        "Does the Sales team have clean, consolidated CRM data — with complete contact records, interaction history, and deal stage data — that could be used to train a revenue prediction model?",
        dq=1, use_case_hint="CRM data quality for AI"),
    row(FN, FN_NAME, DOM, "Piloting", "SAL-P-01",
        "Is AI used for sales forecasting — generating revenue projections with confidence intervals that are more accurate than traditional quota-based forecasting?",
        tr=1, dq=1, vr=1, use_case_hint="AI revenue forecasting"),
    row(FN, FN_NAME, DOM, "Piloting", "SAL-P-02",
        "Does the Sales function use AI for next-best-action recommendations — guiding sellers on the optimal outreach sequence, messaging, and timing for each prospect?",
        tr=1, vr=1, use_case_hint="AI next-best-action sales"),
    row(FN, FN_NAME, DOM, "Piloting", "SAL-P-03",
        "Is AI used for pricing optimisation — dynamically adjusting pricing within guardrails based on demand signals, competitive intelligence, and customer willingness-to-pay models?",
        tr=1, dq=1, vr=1, ge=1, use_case_hint="AI dynamic pricing"),
    row(FN, FN_NAME, DOM, "Scaling", "SAL-S-01",
        "Is AI embedded in the sales process end-to-end — from AI-assisted prospecting to AI-generated proposals, AI-coached calls, and AI-predicted churn risk for existing customers?",
        tr=1, ts=1, vr=1, use_case_hint="End-to-end AI sales process"),
]
ALL_ROWS += cc(FN, FN_NAME, DOM,
    "Does the Sales team have reliable, accessible, clean CRM and sales transaction data — with sufficient historical deal data (volume, velocity, win/loss, customer firmographics) to train an accurate AI pipeline forecasting or lead-scoring model? Is data quality monitored, and are known data gaps on a remediation roadmap?",
    "Have Sales team members been trained on how to interpret and critically evaluate AI-generated lead scores and revenue forecasts — understanding confidence levels, when to override the model, and how to feed deal outcome data back to improve model accuracy?")

# ─── MARKETING & CX ──────────────────────────────────────────────────────────
FN, FN_NAME, DOM = "marketing", "Marketing & CX", "Business"
ALL_ROWS += [
    row(FN, FN_NAME, DOM, "Exploring", "MKT-E-01",
        "Has the Marketing function explored AI tools for content generation, audience segmentation, personalisation, or campaign optimisation?",
        cr=1, vr=1, use_case_hint="Marketing AI exploration"),
    row(FN, FN_NAME, DOM, "Exploring", "MKT-E-02",
        "Does Marketing have unified customer data — combining behavioural, transactional, and demographic data in a Customer Data Platform (CDP) or equivalent — accessible for AI-powered segmentation?",
        dq=1, tr=1, use_case_hint="CDP for AI personalisation"),
    row(FN, FN_NAME, DOM, "Piloting", "MKT-P-01",
        "Is AI used for personalisation at scale — delivering individualised content, offers, or product recommendations across email, web, and mobile based on real-time behavioural signals?",
        tr=1, dq=1, vr=1, use_case_hint="AI personalisation engine"),
    row(FN, FN_NAME, DOM, "Piloting", "MKT-P-02",
        "Is generative AI used for content creation — with guardrails ensuring brand compliance, factual accuracy review, and human approval before publication?",
        tr=1, ge=1, ts=1, vr=1, use_case_hint="Guardrailed GenAI content", reg_ref="EU AI Act"),
    row(FN, FN_NAME, DOM, "Piloting", "MKT-P-03",
        "Is AI used for customer churn prediction — identifying at-risk customers before they disengage and triggering automated retention workflows?",
        tr=1, dq=1, vr=1, use_case_hint="AI churn prediction"),
    row(FN, FN_NAME, DOM, "Scaling", "MKT-S-01",
        "Is AI embedded in the full marketing funnel — from AI-optimised media buying and attribution modelling to AI-generated creative variants and real-time CX personalisation?",
        tr=1, ts=1, dq=1, vr=1, use_case_hint="Full-funnel AI marketing"),
]
ALL_ROWS += cc(FN, FN_NAME, DOM,
    "Does the Marketing team have reliable, accessible, clean customer data — including behavioural event data, purchase history, and consented demographic data — sufficient to train a customer segmentation or personalisation AI model? Is data quality monitored, and are known data gaps on a remediation roadmap?",
    "Have Marketing team members been trained on the responsible use of generative AI for content — including how to identify AI hallucinations, check factual claims, avoid copyright infringement, and comply with FTC/ASA disclosure requirements for AI-generated advertising content?")

# ─── LEGAL & COMPLIANCE ───────────────────────────────────────────────────────
FN, FN_NAME, DOM = "legal", "Legal & Compliance", "Business"
ALL_ROWS += [
    row(FN, FN_NAME, DOM, "Exploring", "LEG-E-01",
        "Is the Legal & Compliance function aware of the EU AI Act, UK AI White Paper, and other applicable AI regulations — and has a preliminary regulatory impact assessment been completed?",
        ge=1, cr=1, use_case_hint="AI regulatory awareness", reg_ref="EU AI Act, UK AI White Paper"),
    row(FN, FN_NAME, DOM, "Exploring", "LEG-E-02",
        "Has Legal reviewed any AI tools currently used by the organisation — assessing vendor contracts for IP, data ownership, confidentiality, and liability provisions related to AI outputs?",
        ge=1, use_case_hint="AI vendor legal review", reg_ref="GDPR, EU AI Act Art.28"),
    row(FN, FN_NAME, DOM, "Piloting", "LEG-P-01",
        "Is AI used for contract review, clause extraction, or obligation tracking — with qualified lawyers reviewing all AI-generated legal analysis before reliance?",
        tr=1, vr=1, ge=1, ts=1, use_case_hint="AI contract review", reg_ref="SRA, FCA"),
    row(FN, FN_NAME, DOM, "Piloting", "LEG-P-02",
        "Has the organisation established an AI governance committee or AI review board — with Legal & Compliance representation — that approves all material AI deployments before go-live?",
        ge=1, cr=1, use_case_hint="AI governance committee", reg_ref="EU AI Act Art.9"),
    row(FN, FN_NAME, DOM, "Scaling", "LEG-S-01",
        "Does Legal & Compliance have a continuous AI regulatory monitoring process — tracking legislative developments across all operating jurisdictions and triggering impact assessments when relevant regulations change?",
        ge=1, cr=1, use_case_hint="AI regulatory monitoring", reg_ref="EU AI Act, GDPR, DPDPA"),
]
ALL_ROWS += cc(FN, FN_NAME, DOM,
    "Does the Legal function have reliable, accessible, structured data on contracts, obligations, regulatory filings, and compliance records — sufficient to train or fine-tune a contract intelligence or compliance monitoring AI model? Is data quality monitored, and are known data gaps on a remediation roadmap?",
    "Have legal and compliance staff been trained on the limitations of AI contract review tools — including how to identify when AI has missed a material clause, mischaracterised a legal obligation, or generated legally incorrect analysis — and is there a mandatory human review protocol before AI-assisted legal outputs are relied upon?")

# ─── PROCUREMENT & SOURCING ───────────────────────────────────────────────────
FN, FN_NAME, DOM = "procurement", "Procurement & Sourcing", "Business"
ALL_ROWS += [
    row(FN, FN_NAME, DOM, "Exploring", "PRO-E-01",
        "Has the Procurement function explored AI for spend analysis, supplier risk scoring, or contract management automation?",
        cr=1, vr=1, use_case_hint="AI procurement exploration"),
    row(FN, FN_NAME, DOM, "Exploring", "PRO-E-02",
        "Does Procurement have clean, consolidated spend data across all categories, suppliers, and entities — accessible in a format suitable for AI-driven spend analytics?",
        dq=1, use_case_hint="Spend data quality"),
    row(FN, FN_NAME, DOM, "Piloting", "PRO-P-01",
        "Is AI used for supplier risk monitoring — automatically scanning news, financial data, and ESG signals to identify at-risk suppliers before disruption occurs?",
        tr=1, dq=1, vr=1, ge=1, use_case_hint="AI supplier risk intelligence"),
    row(FN, FN_NAME, DOM, "Piloting", "PRO-P-02",
        "Has AI been used to optimise sourcing decisions — identifying alternative suppliers, benchmark pricing against market, or automate RFP scoring?",
        tr=1, vr=1, use_case_hint="AI-assisted sourcing"),
    row(FN, FN_NAME, DOM, "Scaling", "PRO-S-01",
        "Is AI embedded across the procure-to-pay process — from AI-assisted specification writing and supplier matching to automated purchase order generation and invoice exception management?",
        tr=1, ts=1, vr=1, use_case_hint="End-to-end AI procurement"),
]
ALL_ROWS += cc(FN, FN_NAME, DOM,
    "Does the Procurement team have reliable, accessible, clean spend data — including supplier master data, purchase order history, invoice data, and contract terms — sufficient to train an AI spend analytics or supplier risk model? Is data quality monitored, and are known data gaps on a remediation roadmap?",
    "Have Procurement staff been trained on AI-assisted sourcing and contract tools — including how to critically evaluate AI-generated supplier risk scores, question AI recommendations that conflict with market intelligence, and understand what data the AI is using to generate its outputs?")

# ─── STRATEGY & PLANNING ─────────────────────────────────────────────────────
FN, FN_NAME, DOM = "strategy", "Strategy & Planning", "Business"
ALL_ROWS += [
    row(FN, FN_NAME, DOM, "Exploring", "STR-E-01",
        "Has the Strategy function used AI for market intelligence, competitive analysis, or scenario modelling to inform strategic decisions?",
        cr=1, vr=1, use_case_hint="AI for strategic intelligence"),
    row(FN, FN_NAME, DOM, "Exploring", "STR-E-02",
        "Is AI readiness explicitly included as a strategic priority in the organisation's 3–5 year strategy — with named KPIs, investment commitments, and board accountability?",
        ge=1, cr=1, vr=1, use_case_hint="AI as strategic priority"),
    row(FN, FN_NAME, DOM, "Piloting", "STR-P-01",
        "Is AI used for strategic scenario planning — modelling multiple macroeconomic, competitive, and technological scenarios and their financial implications in near real-time?",
        tr=1, dq=1, vr=1, use_case_hint="AI scenario planning"),
    row(FN, FN_NAME, DOM, "Scaling", "STR-S-01",
        "Is AI embedded in the strategic planning cycle — with AI-generated market signals and competitive intelligence routinely informing annual planning, capital allocation decisions, and M&A screening?",
        tr=1, ts=1, vr=1, dq=1, use_case_hint="AI-driven strategic planning"),
]
ALL_ROWS += cc(FN, FN_NAME, DOM,
    "Does the Strategy function have reliable, accessible, clean data — including internal performance data, market data subscriptions, and competitive intelligence — sufficient to train an AI-assisted scenario planning or strategy analytics model? Is data quality monitored, and are known data gaps on a remediation roadmap?",
    "Have Strategy and Planning professionals been trained on AI-assisted intelligence tools — including how to critically evaluate AI-generated market analyses, identify when AI is extrapolating beyond its training data, and maintain strategic judgement in the face of AI recommendations?")

# ─── RISK MANAGEMENT ─────────────────────────────────────────────────────────
FN, FN_NAME, DOM = "risk", "Risk Management", "Business"
ALL_ROWS += [
    row(FN, FN_NAME, DOM, "Exploring", "RSK-E-01",
        "Has the Risk Management function assessed the AI-specific risks facing the organisation — including model risk, data poisoning, algorithmic bias, AI vendor concentration, and regulatory non-compliance?",
        ge=1, cr=1, use_case_hint="AI risk framework", reg_ref="EU AI Act Art.9"),
    row(FN, FN_NAME, DOM, "Exploring", "RSK-E-02",
        "Is AI included in the organisation's risk register — with specific AI risks mapped to likelihood, impact, and mitigation owners?",
        ge=1, use_case_hint="AI in enterprise risk register"),
    row(FN, FN_NAME, DOM, "Piloting", "RSK-P-01",
        "Is AI used for risk monitoring and early warning — analysing structured and unstructured data to identify emerging risks before they materialise as losses or incidents?",
        tr=1, dq=1, vr=1, ge=1, use_case_hint="AI risk early warning"),
    row(FN, FN_NAME, DOM, "Piloting", "RSK-P-02",
        "Has a model risk management framework been established — covering validation, monitoring, and decommissioning of all AI models used in material risk or decision-making processes?",
        ge=1, tr=1, use_case_hint="Model risk management framework", reg_ref="SS1/23, SR 11-7"),
    row(FN, FN_NAME, DOM, "Scaling", "RSK-S-01",
        "Is AI risk quantified and reported to the board in monetary terms — with stress-tested scenarios showing the financial exposure of AI model failure, bias incidents, or regulatory sanction?",
        ge=1, vr=1, use_case_hint="AI risk quantification", reg_ref="EU AI Act Art.9"),
]
ALL_ROWS += cc(FN, FN_NAME, DOM,
    "Does the Risk function have reliable, accessible, clean risk event data, KRI time-series, and loss data sufficient to train an AI risk prediction or early-warning model? Is data quality monitored, and are known data gaps on a remediation roadmap?",
    "Have risk professionals been trained to assess and challenge AI model outputs — understanding model risk, overfitting, distributional shift, and when a model's historical training data may no longer reflect current risk conditions?")

# ─── CUSTOMER SERVICE ─────────────────────────────────────────────────────────
FN, FN_NAME, DOM = "customerservice", "Customer Service", "Business"
ALL_ROWS += [
    row(FN, FN_NAME, DOM, "Exploring", "CS-E-01",
        "Has the Customer Service function deployed or piloted any AI — such as chatbots, virtual assistants, AI-assisted agent tools, or automated contact classification?",
        cr=1, vr=1, use_case_hint="AI customer service pilot"),
    row(FN, FN_NAME, DOM, "Exploring", "CS-E-02",
        "Does the Customer Service function have clean, structured interaction data — including chat logs, call transcripts, resolution codes, and CSAT scores — accessible for AI model training?",
        dq=1, use_case_hint="CX interaction data for AI"),
    row(FN, FN_NAME, DOM, "Piloting", "CS-P-01",
        "Is a generative AI-powered conversational agent deployed — resolving customer enquiries autonomously at tier-0 with human escalation for complex or high-value cases?",
        tr=1, vr=1, ge=1, use_case_hint="GenAI conversational agent", reg_ref="EU AI Act"),
    row(FN, FN_NAME, DOM, "Piloting", "CS-P-02",
        "Is AI used for agent-assist — providing real-time suggested responses, knowledge article retrieval, and next-best-action prompts to human agents during live interactions?",
        tr=1, ts=1, vr=1, use_case_hint="AI agent-assist tools"),
    row(FN, FN_NAME, DOM, "Piloting", "CS-P-03",
        "Is sentiment analysis and intent detection applied to all customer interactions — automatically flagging distressed customers, escalation signals, and regulatory vulnerability indicators?",
        tr=1, dq=1, vr=1, ge=1, use_case_hint="AI sentiment & intent detection", reg_ref="FCA Consumer Duty"),
    row(FN, FN_NAME, DOM, "Scaling", "CS-S-01",
        "Is AI deeply embedded in customer service operations — with AI-driven quality assurance covering 100% of interactions, AI-predicted contact volumes driving workforce planning, and AI-personalised customer journeys?",
        tr=1, ts=1, dq=1, vr=1, use_case_hint="AI-native CX operations"),
]
ALL_ROWS += cc(FN, FN_NAME, DOM,
    "Does the Customer Service team have reliable, accessible, clean interaction data — including labelled intent categories, resolution outcomes, customer satisfaction scores, and agent performance data — sufficient to train an AI contact classification or resolution prediction model? Is data quality monitored, and are known data gaps on a remediation roadmap?",
    "Have customer service staff been trained on working alongside AI tools — including how to identify when an AI-generated response is factually incorrect or inappropriate, override AI recommendations for vulnerable customers, and provide feedback to improve AI model quality?")

# ─── CORPORATE GOVERNANCE ─────────────────────────────────────────────────────
FN, FN_NAME, DOM = "corpgovernance", "Corporate Governance", "Business"
ALL_ROWS += [
    row(FN, FN_NAME, DOM, "Exploring", "CG-E-01",
        "Has the Board received a briefing on AI — covering the organisation's AI activity, AI-related risks (regulatory, reputational, operational), and the board's governance obligations under emerging AI legislation?",
        ge=1, cr=1, use_case_hint="Board AI governance briefing", reg_ref="EU AI Act Art.9, UK AI White Paper"),
    row(FN, FN_NAME, DOM, "Exploring", "CG-E-02",
        "Is AI risk included in the audit committee's remit — with internal or external audit coverage of AI model risk, data governance, and AI regulatory compliance?",
        ge=1, use_case_hint="Audit committee AI oversight", reg_ref="EU AI Act Art.9"),
    row(FN, FN_NAME, DOM, "Piloting", "CG-P-01",
        "Has the organisation published or adopted an AI Ethics Policy — covering principles of fairness, transparency, accountability, privacy, and human oversight for all material AI deployments?",
        ge=1, cr=1, use_case_hint="AI ethics policy", reg_ref="EU AI Act Art.9, ISO 42001"),
    row(FN, FN_NAME, DOM, "Piloting", "CG-P-02",
        "Is there a formal AI incident reporting process — requiring material AI failures (bias incidents, model errors causing harm, regulatory breaches) to be escalated to the board?",
        ge=1, use_case_hint="AI incident escalation", reg_ref="EU AI Act Art.73"),
    row(FN, FN_NAME, DOM, "Scaling", "CG-S-01",
        "Does the Annual Report include substantive disclosure on AI governance, AI risk management, and AI-related opportunities — going beyond boilerplate to provide material information to investors?",
        ge=1, vr=1, use_case_hint="AI governance disclosure", reg_ref="TCFD, EU AI Act"),
]
ALL_ROWS += cc(FN, FN_NAME, DOM,
    "Does the Governance function have reliable, accessible, structured data on AI deployments, model inventory, risk assessments, and governance decisions — sufficient to provide the board with an accurate, current picture of the organisation's AI risk exposure? Is this data quality monitored, and are known gaps on a remediation roadmap?",
    "Have board members and senior executives been trained on AI governance obligations — including their personal liability under the EU AI Act, the board's role in approving high-risk AI systems, and how to evaluate AI risk management information presented to them?")

# ─── IT & INFRASTRUCTURE ─────────────────────────────────────────────────────
FN, FN_NAME, DOM = "it", "IT & Infrastructure", "Technology"
ALL_ROWS += [
    row(FN, FN_NAME, DOM, "Exploring", "IT-E-01",
        "Does the organisation's IT infrastructure have the compute capacity, GPU availability, and network bandwidth required to run AI model inference at the scale needed for planned use cases?",
        tr=1, use_case_hint="AI infrastructure capacity"),
    row(FN, FN_NAME, DOM, "Exploring", "IT-E-02",
        "Has an AI infrastructure assessment been completed — identifying gaps in storage, compute, networking, and security architecture that would need to be addressed before AI deployment at scale?",
        tr=1, dq=1, use_case_hint="AI infrastructure gap assessment"),
    row(FN, FN_NAME, DOM, "Piloting", "IT-P-01",
        "Is there a cloud AI platform or managed ML platform (e.g. AWS SageMaker, Azure ML, GCP Vertex AI) provisioned and accessible to data science and engineering teams?",
        tr=1, use_case_hint="Cloud AI/ML platform"),
    row(FN, FN_NAME, DOM, "Piloting", "IT-P-02",
        "Are AI workloads running in production secured — with AI-specific access controls, model endpoint security, data-in-transit encryption, and adversarial input monitoring?",
        tr=1, ge=1, use_case_hint="AI workload security", reg_ref="EU AI Act Art.15"),
    row(FN, FN_NAME, DOM, "Piloting", "IT-P-03",
        "Is AIOps in use — applying ML to IT operations data (logs, metrics, traces) to automate anomaly detection, root cause analysis, and incident prediction?",
        tr=1, vr=1, use_case_hint="AIOps for IT operations"),
    row(FN, FN_NAME, DOM, "Scaling", "IT-S-01",
        "Is the IT function running a FinOps practice for AI — optimising GPU/TPU spend, right-sizing AI inference infrastructure, and tracking AI infrastructure cost per AI use case?",
        tr=1, vr=1, use_case_hint="AI FinOps"),
]
ALL_ROWS += cc(FN, FN_NAME, DOM,
    "Does the IT function have reliable, accessible, clean infrastructure telemetry data — including logs, metrics, CMDB records, and incident history — sufficient to train an AIOps or predictive maintenance AI model? Is data quality monitored, and are known data gaps on a remediation roadmap?",
    "Have IT operations staff been trained on the specific security, observability, and governance requirements of AI workloads — including how to monitor AI model endpoints, detect model degradation, respond to AI-specific security incidents, and understand MLOps practices?")

# ─── SOFTWARE DEVELOPMENT ─────────────────────────────────────────────────────
FN, FN_NAME, DOM = "softwaredev", "Software Development", "Technology"
ALL_ROWS += [
    row(FN, FN_NAME, DOM, "Exploring", "SD-E-01",
        "Are software developers in the organisation using AI coding assistants (GitHub Copilot, Cursor, Amazon CodeWhisperer, etc.) — and is their use governed by a policy covering IP, security, and code review requirements?",
        tr=1, ge=1, ts=1, use_case_hint="AI coding assistant governance"),
    row(FN, FN_NAME, DOM, "Exploring", "SD-E-02",
        "Has the engineering team assessed the security implications of AI-generated code — including vulnerability scanning of AI-assisted code before merge and awareness training on common AI coding mistakes?",
        tr=1, ge=1, ts=1, use_case_hint="AI code security review"),
    row(FN, FN_NAME, DOM, "Piloting", "SD-P-01",
        "Is AI used for automated test generation — creating unit tests, integration tests, and edge-case tests automatically from code changes?",
        tr=1, vr=1, use_case_hint="AI automated testing"),
    row(FN, FN_NAME, DOM, "Piloting", "SD-P-02",
        "Does the team have an MLOps pipeline — enabling automated model training, validation, deployment, and rollback for AI models built internally?",
        tr=1, dq=1, use_case_hint="MLOps CI/CD pipeline"),
    row(FN, FN_NAME, DOM, "Scaling", "SD-S-01",
        "Is AI embedded in the full software development lifecycle — from AI-assisted requirements analysis and architecture review to AI-generated documentation and AI-driven production monitoring?",
        tr=1, ts=1, vr=1, use_case_hint="AI-native SDLC"),
]
ALL_ROWS += cc(FN, FN_NAME, DOM,
    "Does the Software Development team have reliable, accessible, clean code and engineering data — including version control history, code quality metrics, test coverage data, and incident post-mortems — sufficient to train an AI defect prediction or developer productivity model? Is data quality monitored, and are known data gaps on a remediation roadmap?",
    "Have software developers been trained on the responsible use of AI coding assistants — including how to critically review AI-generated code for security vulnerabilities, avoid accepting suggestions that introduce licence-incompatible code, and understand when AI-generated code requires expert human review before production deployment?")

# ─── DATA & ANALYTICS ─────────────────────────────────────────────────────────
FN, FN_NAME, DOM = "data", "Data & Analytics", "Technology"
ALL_ROWS += [
    row(FN, FN_NAME, DOM, "Exploring", "DA-E-01",
        "Does the organisation have a functioning data analytics capability — with BI tools, dashboards, and self-service analytics accessible to business functions beyond the data team?",
        dq=1, tr=1, use_case_hint="Analytics capability baseline"),
    row(FN, FN_NAME, DOM, "Exploring", "DA-E-02",
        "Are data scientists or ML engineers employed or contracted within the organisation — with the capability to build, train, and deploy custom ML models?",
        ts=1, tr=1, use_case_hint="Internal ML capability"),
    row(FN, FN_NAME, DOM, "Piloting", "DA-P-01",
        "Has the Data team delivered at least one AI or ML model in production — with a measurable business outcome tracked against a defined success metric?",
        tr=1, dq=1, vr=1, use_case_hint="First production ML model"),
    row(FN, FN_NAME, DOM, "Piloting", "DA-P-02",
        "Is there an experimentation platform — enabling A/B tests, multi-armed bandits, or ML experiments with statistical rigour and automated winning variant deployment?",
        tr=1, dq=1, use_case_hint="ML experimentation platform"),
    row(FN, FN_NAME, DOM, "Scaling", "DA-S-01",
        "Does the Data team operate an internal AI product platform — providing reusable AI capabilities (recommendation engines, NLP APIs, forecasting services) to other business functions via APIs?",
        tr=1, ts=1, vr=1, use_case_hint="Internal AI platform-as-a-service"),
]
ALL_ROWS += cc(FN, FN_NAME, DOM,
    "Does the Data & Analytics function have reliable, accessible, clean data — with documented schemas, data contracts, and quality SLAs — such that models built by the team can be reliably retrained and reproduced? Is data quality monitored continuously, and are known data gaps on a remediation roadmap?",
    "Have data analysts and data scientists been trained on the responsible AI practices specific to their role — including how to detect and mitigate training data bias, perform model explainability analysis, validate models against held-out test sets, and document model cards for governance review?")

# ─── CYBERSECURITY ─────────────────────────────────────────────────────────
FN, FN_NAME, DOM = "cybersecurity", "Cybersecurity", "Technology"
ALL_ROWS += [
    row(FN, FN_NAME, DOM, "Exploring", "CYB-E-01",
        "Has a cybersecurity risk assessment been conducted specifically for AI — covering adversarial attacks, model theft, data poisoning, prompt injection, and AI-generated social engineering threats?",
        ge=1, tr=1, use_case_hint="AI-specific security risk assessment", reg_ref="NIST AI RMF"),
    row(FN, FN_NAME, DOM, "Exploring", "CYB-E-02",
        "Is the Security Operations Centre (SOC) aware of and equipped to detect AI-specific threats — including model extraction attacks, API abuse of LLM endpoints, and deepfake-enabled fraud?",
        tr=1, ts=1, use_case_hint="AI threat detection in SOC"),
    row(FN, FN_NAME, DOM, "Piloting", "CYB-P-01",
        "Is AI used in cybersecurity operations — for threat detection, anomaly identification, security log analysis, or automated incident response triage?",
        tr=1, vr=1, use_case_hint="AI-powered SOC"),
    row(FN, FN_NAME, DOM, "Piloting", "CYB-P-02",
        "Are all production AI models protected by security controls — including input validation, output filtering, rate limiting, and monitoring for adversarial inputs or prompt injection attempts?",
        tr=1, ge=1, use_case_hint="AI model security controls", reg_ref="EU AI Act Art.15"),
    row(FN, FN_NAME, DOM, "Scaling", "CYB-S-01",
        "Is the organisation actively contributing to or monitoring AI security research — tracking LLM jailbreaks, adversarial ML developments, and AI supply chain attacks, and updating defences proactively?",
        tr=1, ge=1, cr=1, use_case_hint="Proactive AI security research"),
]
ALL_ROWS += cc(FN, FN_NAME, DOM,
    "Does the Cybersecurity team have reliable, accessible, clean security telemetry data — including network logs, endpoint events, threat intelligence feeds, and incident records — sufficient to train an AI threat detection or anomaly detection model? Is data quality monitored, and are known data gaps on a remediation roadmap?",
    "Have security analysts and engineers been trained on AI-specific attack vectors — including adversarial machine learning, prompt injection, model inversion attacks, and AI-generated phishing — and do they understand how to test AI systems for these vulnerabilities before production deployment?")

# ─── IT SERVICE MANAGEMENT ────────────────────────────────────────────────────
FN, FN_NAME, DOM = "itsm", "IT Service Management", "Technology"
ALL_ROWS += [
    row(FN, FN_NAME, DOM, "Exploring", "ITSM-E-01",
        "Has the ITSM function explored AI for ticket classification, auto-routing, or knowledge base article suggestion to reduce mean time to resolve (MTTR)?",
        cr=1, vr=1, use_case_hint="AI ticket triage"),
    row(FN, FN_NAME, DOM, "Exploring", "ITSM-E-02",
        "Does the ITSM platform capture structured, clean service data — with consistent incident categorisation, resolution codes, and CI associations — sufficient to train an AI classification model?",
        dq=1, tr=1, use_case_hint="ITSM data quality for AI"),
    row(FN, FN_NAME, DOM, "Piloting", "ITSM-P-01",
        "Is a conversational AI or virtual agent deployed for tier-0 IT support — resolving common requests (password resets, software installation, access provisioning) without human agent involvement?",
        tr=1, vr=1, use_case_hint="AI tier-0 IT support"),
    row(FN, FN_NAME, DOM, "Piloting", "ITSM-P-02",
        "Is AI used for predictive incident management — detecting patterns in monitoring data that predict incidents before users are impacted, enabling proactive resolution?",
        tr=1, dq=1, vr=1, use_case_hint="Predictive incident management"),
    row(FN, FN_NAME, DOM, "Scaling", "ITSM-S-01",
        "Is the ITSM function operating at AI-Native maturity — with AI-driven change risk scoring, AI-assisted problem management root cause analysis, and AI-predicted service demand driving capacity planning?",
        tr=1, ts=1, vr=1, dq=1, use_case_hint="AI-native ITSM operations"),
]
ALL_ROWS += cc(FN, FN_NAME, DOM,
    "Does the ITSM team have reliable, accessible, clean service management data — including correctly categorised incidents, problems, changes, and known errors in a structured CMDB — sufficient to train an AI classification or prediction model? Is data quality monitored, and are known data gaps on a remediation roadmap?",
    "Have ITSM staff been trained on working alongside AI tools for ticket triage and knowledge management — including how to identify when AI routing decisions are incorrect, provide feedback to improve the model, and maintain human oversight for high-impact change approvals?")

# ─── CLOUD & PLATFORM ────────────────────────────────────────────────────────
FN, FN_NAME, DOM = "cloud", "Cloud & Platform", "Technology"
ALL_ROWS += [
    row(FN, FN_NAME, DOM, "Exploring", "CLD-E-01",
        "Has the organisation assessed which cloud platform (AWS, Azure, GCP, or multi-cloud) is best positioned to host its AI workloads — considering data residency, AI service availability, and cost?",
        tr=1, ge=1, use_case_hint="Cloud AI platform selection"),
    row(FN, FN_NAME, DOM, "Piloting", "CLD-P-01",
        "Are cloud-native AI services (AWS Bedrock, Azure OpenAI, GCP Vertex AI) being consumed by development teams — with usage governed by policies on data handling, cost controls, and approved models?",
        tr=1, ge=1, vr=1, use_case_hint="Governed cloud AI service consumption"),
    row(FN, FN_NAME, DOM, "Piloting", "CLD-P-02",
        "Is there a cloud cost management practice for AI — with tagging of AI workloads, budget alerts for AI spend, and regular optimisation reviews to prevent AI cost overruns?",
        tr=1, vr=1, use_case_hint="AI cloud cost management"),
    row(FN, FN_NAME, DOM, "Scaling", "CLD-S-01",
        "Is the organisation running a multi-cloud AI strategy — deploying AI workloads across providers to optimise cost, resilience, and access to best-in-class AI capabilities?",
        tr=1, vr=1, use_case_hint="Multi-cloud AI strategy"),
]
ALL_ROWS += cc(FN, FN_NAME, DOM,
    "Does the Cloud & Platform team have reliable, accessible, clean infrastructure and cost data — including cloud usage telemetry, resource tagging, and billing records — sufficient to train an AI cost optimisation or capacity forecasting model? Is data quality monitored, and are known data gaps on a remediation roadmap?",
    "Have cloud engineers been trained on the AI-specific services, security configurations, and cost models of their cloud provider(s) — including responsible use of foundation model APIs, data residency configuration, and how to implement guardrails for generative AI applications?")

# ─── NETWORK & COMMUNICATIONS ─────────────────────────────────────────────────
FN, FN_NAME, DOM = "network", "Network & Communications", "Technology"
ALL_ROWS += [
    row(FN, FN_NAME, DOM, "Exploring", "NET-E-01",
        "Has the network infrastructure team assessed whether current bandwidth, latency, and edge compute capacity is sufficient for real-time AI inference workloads at the locations where AI will be deployed?",
        tr=1, use_case_hint="Network AI readiness assessment"),
    row(FN, FN_NAME, DOM, "Piloting", "NET-P-01",
        "Is AI used for network operations — automating anomaly detection in traffic patterns, predicting network congestion, or optimising routing decisions in real time?",
        tr=1, vr=1, use_case_hint="AI network operations"),
    row(FN, FN_NAME, DOM, "Scaling", "NET-S-01",
        "Is the organisation deploying edge AI — running AI inference at the network edge (IoT gateways, retail endpoints, industrial controllers) to enable real-time AI decisions without cloud round-trips?",
        tr=1, vr=1, use_case_hint="Edge AI deployment"),
]
ALL_ROWS += cc(FN, FN_NAME, DOM,
    "Does the Network team have reliable, accessible, clean network telemetry data — including flow data, device logs, and performance metrics — sufficient to train an AI network anomaly detection or predictive maintenance model? Is data quality monitored, and are known data gaps on a remediation roadmap?",
    "Have network engineers been trained on the specific requirements of supporting AI workloads — including low-latency networking for AI inference, SD-WAN configuration for AI traffic prioritisation, and the network security implications of AI model endpoints exposed on the network?")

# ─── DIGITAL WORKPLACE ────────────────────────────────────────────────────────
FN, FN_NAME, DOM = "digitalworkplace", "Digital Workplace", "Technology"
ALL_ROWS += [
    row(FN, FN_NAME, DOM, "Exploring", "DW-E-01",
        "Has the Digital Workplace function assessed which productivity AI tools (Microsoft Copilot, Google Workspace AI, Notion AI, etc.) are appropriate for the organisation — and established an approved tools list with security and data classification requirements?",
        ge=1, ts=1, cr=1, use_case_hint="Approved workplace AI tools list"),
    row(FN, FN_NAME, DOM, "Piloting", "DW-P-01",
        "Is Microsoft 365 Copilot or an equivalent AI productivity tool deployed and actively used by employees — with usage metrics tracked and productivity impact measured?",
        tr=1, ts=1, vr=1, use_case_hint="AI productivity tool deployment"),
    row(FN, FN_NAME, DOM, "Piloting", "DW-P-02",
        "Is there a process for employees to request and access new AI tools — including a lightweight security review, data classification check, and manager approval workflow?",
        ge=1, cr=1, use_case_hint="AI tool request process"),
    row(FN, FN_NAME, DOM, "Scaling", "DW-S-01",
        "Is AI deeply embedded in the digital workplace — with AI-powered knowledge management, AI-assisted meeting summarisation, AI-driven workflow automation, and AI-personalised information discovery available to all employees?",
        tr=1, ts=1, vr=1, use_case_hint="AI-native digital workplace"),
]
ALL_ROWS += cc(FN, FN_NAME, DOM,
    "Does the Digital Workplace team have reliable, accessible, clean usage data — including adoption metrics, productivity measurements, and employee feedback on AI tools — sufficient to evaluate AI productivity impact and train a recommendation model for tool adoption? Is data quality monitored, and are known data gaps on a remediation roadmap?",
    "Have employees been trained on the approved AI productivity tools — going beyond basic feature training to include critical use of AI outputs, understanding what the AI tool can and cannot do reliably, data classification rules for AI tool use, and how to report concerns about AI outputs?")

# ─── SUPPLY CHAIN & LOGISTICS ─────────────────────────────────────────────────
FN, FN_NAME, DOM = "supplychain", "Supply Chain & Logistics", "Operations"
ALL_ROWS += [
    row(FN, FN_NAME, DOM, "Exploring", "SC-E-01",
        "Has the Supply Chain function explored AI for demand forecasting, inventory optimisation, or logistics route planning?",
        cr=1, vr=1, use_case_hint="AI supply chain exploration"),
    row(FN, FN_NAME, DOM, "Exploring", "SC-E-02",
        "Does the Supply Chain team have clean, consolidated data on demand history, inventory levels, lead times, and supplier performance — accessible in a format suitable for AI model training?",
        dq=1, use_case_hint="Supply chain data quality"),
    row(FN, FN_NAME, DOM, "Piloting", "SC-P-01",
        "Is AI-driven demand forecasting deployed — generating probabilistic forecasts by SKU, location, and time horizon that are more accurate than statistical baselines?",
        tr=1, dq=1, vr=1, use_case_hint="AI demand forecasting"),
    row(FN, FN_NAME, DOM, "Piloting", "SC-P-02",
        "Is AI used for supply chain disruption prediction — monitoring supplier risk signals, geopolitical events, and logistics data to anticipate disruptions 4–8 weeks ahead?",
        tr=1, dq=1, vr=1, ge=1, use_case_hint="AI disruption prediction"),
    row(FN, FN_NAME, DOM, "Scaling", "SC-S-01",
        "Is the supply chain operating as an AI-driven autonomous network — with AI optimising replenishment, routing, and supplier allocation in near-real-time across the end-to-end value chain?",
        tr=1, ts=1, dq=1, vr=1, use_case_hint="Autonomous AI supply chain"),
]
ALL_ROWS += cc(FN, FN_NAME, DOM,
    "Does the Supply Chain team have reliable, accessible, clean data — including multi-year demand history, inventory snapshots, supplier lead times, and logistics performance data — sufficient to train an AI demand forecasting or inventory optimisation model? Is data quality monitored, and are known data gaps on a remediation roadmap?",
    "Have supply chain planners been trained to interpret AI demand forecasts — understanding confidence intervals, when to override the model, how to identify when external events (promotions, new product launches) invalidate the forecast, and how to feed exceptions back to improve model accuracy?")

# ─── MANUFACTURING & PRODUCTION ───────────────────────────────────────────────
FN, FN_NAME, DOM = "manufacturing", "Manufacturing & Production", "Operations"
ALL_ROWS += [
    row(FN, FN_NAME, DOM, "Exploring", "MFG-E-01",
        "Has the Manufacturing function explored AI for predictive maintenance, quality control, or production optimisation on the factory floor?",
        cr=1, vr=1, use_case_hint="AI manufacturing exploration"),
    row(FN, FN_NAME, DOM, "Exploring", "MFG-E-02",
        "Does the Manufacturing function have sensor data, machine telemetry, and quality inspection data in a format and volume sufficient to train predictive maintenance or defect detection AI models?",
        dq=1, tr=1, use_case_hint="Manufacturing IoT data readiness"),
    row(FN, FN_NAME, DOM, "Piloting", "MFG-P-01",
        "Is AI-powered predictive maintenance deployed — using sensor data to predict equipment failure 24–72 hours ahead, enabling planned maintenance before unplanned downtime occurs?",
        tr=1, dq=1, vr=1, use_case_hint="Predictive maintenance AI"),
    row(FN, FN_NAME, DOM, "Piloting", "MFG-P-02",
        "Is computer vision used for automated quality inspection — detecting defects, foreign objects, or specification deviations on the production line faster and more accurately than manual inspection?",
        tr=1, dq=1, vr=1, use_case_hint="Computer vision QC"),
    row(FN, FN_NAME, DOM, "Scaling", "MFG-S-01",
        "Is AI optimising production scheduling, yield, and energy consumption in real time — with AI models running on edge devices at the factory floor and feeding decisions back into ERP/MES systems?",
        tr=1, ts=1, dq=1, vr=1, use_case_hint="AI production optimisation"),
]
ALL_ROWS += cc(FN, FN_NAME, DOM,
    "Does the Manufacturing team have reliable, accessible, clean operational data — including machine sensor readings, maintenance logs, quality inspection records, and production output data — sufficient to train a predictive maintenance or defect detection AI model? Is data quality monitored, and are known data gaps on a remediation roadmap?",
    "Have manufacturing engineers and operators been trained to work alongside AI quality and maintenance tools — understanding what the AI model is monitoring, how to interpret AI alerts, when to override AI-driven maintenance recommendations, and how to report sensor anomalies that may affect model accuracy?")

# ─── QUALITY ASSURANCE ────────────────────────────────────────────────────────
FN, FN_NAME, DOM = "qualityassurance", "Quality Assurance", "Operations"
ALL_ROWS += [
    row(FN, FN_NAME, DOM, "Exploring", "QA-E-01",
        "Has the QA function explored AI for defect prediction, audit sampling optimisation, or quality management system automation?",
        cr=1, vr=1, use_case_hint="AI QA exploration"),
    row(FN, FN_NAME, DOM, "Piloting", "QA-P-01",
        "Is AI used to prioritise audit and inspection activities — using risk-based models to direct QA resource to highest-risk products, processes, or suppliers?",
        tr=1, dq=1, vr=1, use_case_hint="AI risk-based QA prioritisation"),
    row(FN, FN_NAME, DOM, "Piloting", "QA-P-02",
        "Is AI used to analyse non-conformance and customer complaint data — identifying systemic root causes and predicting future quality failures before they reach customers?",
        tr=1, dq=1, vr=1, use_case_hint="AI quality failure prediction"),
    row(FN, FN_NAME, DOM, "Scaling", "QA-S-01",
        "Is quality control AI-native — with AI systems monitoring 100% of production output for defects, generating real-time quality dashboards, and automatically adjusting process parameters to maintain specification?",
        tr=1, ts=1, dq=1, vr=1, use_case_hint="AI-native quality control"),
]
ALL_ROWS += cc(FN, FN_NAME, DOM,
    "Does the QA team have reliable, accessible, clean quality data — including non-conformance records, inspection results, complaint data, and supplier quality scores — sufficient to train an AI defect prediction or risk-based inspection model? Is data quality monitored, and are known data gaps on a remediation roadmap?",
    "Have QA professionals been trained on AI-assisted inspection and quality analytics tools — including how to validate AI defect detection accuracy, understand false positive/negative rates, and maintain human oversight for safety-critical quality decisions?")

# ─── FACILITIES & PROPERTY ────────────────────────────────────────────────────
FN, FN_NAME, DOM = "facilities", "Facilities & Property", "Operations"
ALL_ROWS += [
    row(FN, FN_NAME, DOM, "Exploring", "FAC-E-01",
        "Has the Facilities function explored AI or smart building technology for energy optimisation, space utilisation analysis, or predictive maintenance of building systems?",
        cr=1, vr=1, use_case_hint="Smart building AI exploration"),
    row(FN, FN_NAME, DOM, "Piloting", "FAC-P-01",
        "Is AI used to optimise energy consumption across the property portfolio — using occupancy data, weather forecasts, and utility pricing to minimise cost and carbon footprint?",
        tr=1, dq=1, vr=1, use_case_hint="AI building energy optimisation"),
    row(FN, FN_NAME, DOM, "Scaling", "FAC-S-01",
        "Is the property portfolio managed by an AI-driven facilities intelligence platform — providing real-time space utilisation, predictive maintenance scheduling, and energy optimisation across all sites?",
        tr=1, ts=1, vr=1, use_case_hint="AI facilities management platform"),
]
ALL_ROWS += cc(FN, FN_NAME, DOM,
    "Does the Facilities team have reliable, accessible, clean building operations data — including BMS sensor data, energy consumption records, maintenance history, and space utilisation data — sufficient to train a building energy optimisation or predictive maintenance AI model? Is data quality monitored, and are known data gaps on a remediation roadmap?",
    "Have facilities managers been trained on AI-powered building management systems — including how to interpret AI-generated maintenance predictions, override automated system adjustments when operational context requires it, and feed actual outcomes back to improve model accuracy?")

# ─── HEALTH & SAFETY ──────────────────────────────────────────────────────────
FN, FN_NAME, DOM = "healthsafety", "Health & Safety", "Operations"
ALL_ROWS += [
    row(FN, FN_NAME, DOM, "Exploring", "HS-E-01",
        "Has the Health & Safety function explored AI for incident prediction, safety observation analysis, or near-miss pattern detection?",
        cr=1, ge=1, use_case_hint="AI H&S risk prediction"),
    row(FN, FN_NAME, DOM, "Piloting", "HS-P-01",
        "Is AI used to analyse safety incident data — identifying leading indicators, predicting high-risk areas or activities, and directing safety interventions before incidents occur?",
        tr=1, dq=1, vr=1, ge=1, use_case_hint="AI safety incident prediction", reg_ref="ISO 45001"),
    row(FN, FN_NAME, DOM, "Scaling", "HS-S-01",
        "Is computer vision deployed for real-time safety monitoring — detecting PPE compliance, unsafe behaviours, or hazardous conditions on site and triggering immediate alerts?",
        tr=1, ge=1, vr=1, use_case_hint="Computer vision safety monitoring", reg_ref="EU AI Act Annex III"),
]
ALL_ROWS += cc(FN, FN_NAME, DOM,
    "Does the H&S team have reliable, accessible, clean safety data — including structured incident reports, near-miss records, risk assessments, and safety observation data — sufficient to train an AI incident prediction or safety risk model? Is data quality monitored, and are known data gaps on a remediation roadmap?",
    "Have H&S professionals been trained on AI safety monitoring tools — including the ethical and legal implications of AI surveillance in the workplace, how to ensure AI safety tools do not have disproportionate impact on workers, and how to maintain human judgement as the final decision-maker for safety-critical actions?")

# ─── ENVIRONMENTAL & SUSTAINABILITY ───────────────────────────────────────────
FN, FN_NAME, DOM = "environmental", "Environmental & Sustainability", "Operations"
ALL_ROWS += [
    row(FN, FN_NAME, DOM, "Exploring", "ENV-E-01",
        "Has the Sustainability function assessed the environmental impact of the organisation's AI activities — including the energy consumption and carbon footprint of training and running AI models?",
        ge=1, cr=1, use_case_hint="AI carbon footprint assessment", reg_ref="TCFD, CSRD"),
    row(FN, FN_NAME, DOM, "Piloting", "ENV-P-01",
        "Is AI used for sustainability reporting — automating data collection from operational systems, calculating Scope 1, 2, and 3 emissions, and generating regulatory disclosures?",
        tr=1, dq=1, vr=1, ge=1, use_case_hint="AI sustainability reporting", reg_ref="CSRD, TCFD"),
    row(FN, FN_NAME, DOM, "Piloting", "ENV-P-02",
        "Is AI used to optimise energy, water, waste, and resource consumption across operations — with AI models identifying efficiency opportunities and tracking improvement against targets?",
        tr=1, dq=1, vr=1, use_case_hint="AI resource optimisation"),
    row(FN, FN_NAME, DOM, "Scaling", "ENV-S-01",
        "Has the organisation committed to and implemented a sustainable AI policy — covering model efficiency requirements, preference for renewable-energy data centres, and AI carbon accounting in the sustainability report?",
        ge=1, cr=1, vr=1, use_case_hint="Sustainable AI policy", reg_ref="CSRD, EU AI Act"),
]
ALL_ROWS += cc(FN, FN_NAME, DOM,
    "Does the Sustainability team have reliable, accessible, clean environmental data — including energy consumption records, Scope 1/2/3 emissions data, waste and water data, and supply chain sustainability metrics — sufficient to train an AI emissions forecasting or resource optimisation model? Is data quality monitored, and are known data gaps on a remediation roadmap?",
    "Have sustainability professionals been trained on AI-assisted ESG reporting and environmental optimisation tools — including how to validate AI-calculated emissions figures, identify data quality issues that affect AI outputs, and ensure AI-generated sustainability disclosures meet regulatory accuracy requirements?")

# ─── FIELD SERVICE ────────────────────────────────────────────────────────────
FN, FN_NAME, DOM = "fieldservice", "Field Service", "Operations"
ALL_ROWS += [
    row(FN, FN_NAME, DOM, "Exploring", "FS-E-01",
        "Has the Field Service function explored AI for engineer scheduling optimisation, parts prediction, or remote AI-assisted diagnosis before dispatching an engineer?",
        cr=1, vr=1, use_case_hint="AI field service exploration"),
    row(FN, FN_NAME, DOM, "Piloting", "FS-P-01",
        "Is AI used for intelligent field scheduling — optimising engineer dispatch based on skill, location, availability, traffic, and job priority to maximise first-time fix rates?",
        tr=1, dq=1, vr=1, use_case_hint="AI field scheduling optimisation"),
    row(FN, FN_NAME, DOM, "Piloting", "FS-P-02",
        "Is predictive parts management in use — AI models forecasting which parts will be needed on each job and pre-positioning inventory to reduce van stock-outs and repeat visits?",
        tr=1, dq=1, vr=1, use_case_hint="AI predictive parts management"),
    row(FN, FN_NAME, DOM, "Scaling", "FS-S-01",
        "Is the field service organisation AI-native — with AI-driven remote diagnostics, AR-assisted repair guidance, predictive failure prevention, and autonomous scheduling delivering measurable improvement in first-time fix rate and customer satisfaction?",
        tr=1, ts=1, dq=1, vr=1, use_case_hint="AI-native field service"),
]
ALL_ROWS += cc(FN, FN_NAME, DOM,
    "Does the Field Service team have reliable, accessible, clean job data — including job history, engineer skills records, parts usage data, first-time fix outcomes, and asset service records — sufficient to train an AI scheduling or predictive maintenance model? Is data quality monitored, and are known data gaps on a remediation roadmap?",
    "Have field service engineers been trained on AI-assisted diagnostic and scheduling tools — including how to interpret AI-generated job recommendations, override automated scheduling when on-site context requires it, and report job outcomes accurately to improve AI model performance?")

# ═══════════════════════════════════════════════════════════════════════════════
# BUILD XLSX
# ═══════════════════════════════════════════════════════════════════════════════
def build_xlsx(rows, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Questions"

    # Header row styling
    header_fill = PatternFill("solid", fgColor="1E293B")
    header_font = Font(bold=True, color="F8FAFC", size=10)

    for col_idx, col_name in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, r in enumerate(rows, 2):
        for col_idx, col_name in enumerate(COLS, 1):
            ws.cell(row=row_idx, column=col_idx, value=r.get(col_name, ""))

    # Column widths
    col_widths = {
        "function_id": 18, "function_name": 28, "domain": 14,
        "competency_level": 14, "question_id": 14,
        "question_text": 80, "answer_type": 10,
        "branch_trigger": 12, "follow_up_text": 50,
        "dim_DQ": 7, "dim_TR": 7, "dim_TS": 7,
        "dim_GE": 7, "dim_CR": 7, "dim_VR": 7,
        "points_yes": 9, "points_partial": 10, "points_no": 8,
        "industry_overlay": 14, "ai_use_case_hint": 40, "regulatory_ref": 30,
    }
    for col_idx, col_name in enumerate(COLS, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = col_widths.get(col_name, 12)

    ws.freeze_panes = "A2"
    wb.save(output_path)
    print(f"Saved {len(rows)} rows -> {output_path}")

# ─── Stats ─────────────────────────────────────────────────────────────────
def print_stats(rows):
    from collections import Counter
    fn_counts = Counter(r["function_id"] for r in rows)
    print(f"\nTotal questions: {len(rows)}")
    print(f"Total functions: {len(fn_counts)}")
    print("\nQuestions per function:")
    for fn, count in sorted(fn_counts.items()):
        print(f"  {fn:<30} {count}")
    dims_used = Counter()
    for r in rows:
        for d in ["dim_DQ","dim_TR","dim_TS","dim_GE","dim_CR","dim_VR"]:
            if r.get(d,0): dims_used[d] += 1
    print("\nDimension coverage:")
    for d, count in sorted(dims_used.items()):
        print(f"  {d}: {count} questions")
    # Verify dim weights sum
    print("\nDIMS weights (should be updated in App.jsx): DQ=0.30, TR=0.18, TS=0.22, GE=0.15, CR=0.08, VR=0.07 -> sum=1.00")

if __name__ == "__main__":
    output_path = os.path.join(os.path.dirname(__file__), "public", "ai-question-bank.xlsx")
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    build_xlsx(ALL_ROWS, output_path)
    print_stats(ALL_ROWS)
